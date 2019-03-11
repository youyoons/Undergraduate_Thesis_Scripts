import openpyxl
import os


#ac_num = '4214124324'

def anonymize_acnum(ac_num):
    filepath = "ac_hash.xlsx"
    
    #We should first check if the accession number exists in the hash table before adding it
    if os.path.isfile("ac_hash.xlsx"):
        exists = False
        
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        for i in range(2,len(sheet['A'])+1):
            if sheet.cell(row = i, column = 1).value == ac_num:
                exists = True
                new_ac_num = sheet.cell(row = i, column = 2).value
                break
                
        if not exists: #enter value into table if it does not exist
            row_count = sheet.max_row
            
            sheet.cell(row = row_count + 1, column = 1, value = str(ac_num))
            sheet.cell(row = row_count + 1, column = 2, value = str(row_count).zfill(7))
            
            wb.save(filepath)
            
            new_ac_num = str(row_count).zfill(7)
        
    else: #Create workbook and add the mapping
        #First, save spreadsheet
        wb = openpyxl.Workbook()
        wb.save(filepath)
        
        #Put in hashed value
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        
        #Insert first row of headings
        sheet.cell(row = 1, column = 1, value = "Original AC Num")
        sheet.cell(row = 1, column = 2, value = 'New AC Num')
        
        #Insert second row, which contains data
        row_count = sheet.max_row
        
        sheet.cell(row = row_count + 1, column = 1, value = str(ac_num))
        sheet.cell(row = row_count + 1, column = 2, value = str(row_count).zfill(7))
        
        
        print(row_count)
        wb.save(filepath)
        
        new_ac_num = str(row_count).zfill(7)
    
    return new_ac_num

#print(anonymize_acnum(ac_num))
