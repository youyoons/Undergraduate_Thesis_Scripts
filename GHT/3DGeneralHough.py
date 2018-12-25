import os
import numpy as np
import math
import matplotlib.pyplot as plt
from collections import defaultdict
from scipy.misc import imread
from skimage.feature import canny
from scipy.ndimage.filters import sobel
from mpl_toolkits.mplot3d import Axes3D
import cv2 as cv
from scipy import signal
from scipy.ndimage.filters import gaussian_filter
import datetime
import openpyxl
try:
    import cPickle
except ImportError:
    import pickle as cPickle


def gradient_orientation(image):
    '''
    Calculate the gradient orientation for edge point in the image
    '''
    #scipy.ndimage.sobel
    dx = sobel(image, axis=0, mode='constant')
    dy = sobel(image, axis=1, mode='constant')
    dz = sobel(image, axis=1, mode='constant')
    
    #For 3D instead of a single gradient value, we need two angles that define a normal vector
    #Phi is the angle between the positive x-axis to the projection of the normal vector the x-y plane (around +z)
    #Psi is the angle between the positive z-axis to the normal vector
    
    phi = np.arctan2(dy ,dx) * 180 / np.pi
    psi = np.arctan2(np.sqrt(dx*dx + dy*dy), dz) * 180 / np.pi
    

    gradient = np.zeros(image.shape)
    
    return phi, psi

def build_r_table(image, origin):
    '''
    Build the R-table from the given shape image and a referenc point
    '''
    
    dx = sobel(image, 0)  # x derivative
    dy = sobel(image, 1)  # y derivative
    dz = sobel(image, 2)  # z derivative
    
    mag = np.sqrt(dx*dx + dy*dy + dz*dz)
    mag_norm = mag/np.max(mag)
    
    #Creating edge array the same size as query image
    edges = np.zeros(image.shape, dtype=bool)
    
    for i in range(edges.shape[0]):
        for j in range(edges.shape[1]):
            for k in range(edges.shape[2]):
                if mag_norm[i,j,k] > 0.3 :# and mag_norm[i,j,k] < 0.9:
                    edges[i,j,k] = True
    

    #Takes (47,40) Edges and calculates the gradients using sobel
    phi, psi = gradient_orientation(edges)
    #print("Phi Dim: ", phi.shape)
    
    r_table = defaultdict(list)
    for (i,j,k),value in np.ndenumerate(edges):
        if value:
            r_table[(int(phi[i,j,k]),int(psi[i,j,k]))].append((origin[0]-i, origin[1]-j, origin[2] - k))
    
    
    return r_table

def canny_edges_3d(grayImage):
    global MIN_CANNY_THRESHOLD
    global MAX_CANNY_THRESHOLD

    MIN_CANNY_THRESHOLD = 20
    MAX_CANNY_THRESHOLD = 100
    
    dim = np.shape(grayImage)
    
    edges_x = np.zeros(grayImage.shape, dtype=bool) 
    edges_y = np.zeros(grayImage.shape, dtype=bool) 
    edges_z = np.zeros(grayImage.shape, dtype=bool) 
    edges = np.zeros(grayImage.shape, dtype=bool) 
    

    for i in range(dim[0]):
        edges_x[i,:,:] = canny(grayImage[i,:,:], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = 0)
   
    for j in range(dim[1]):
        edges_y[:,j,:] = canny(grayImage[:,j,:], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = 0)
        
    for k in range(dim[2]):
        edges_z[:,:,k] = canny(grayImage[:,:,k], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = 0)
    
    
   # edges = canny(grayImage, low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD)
    for i in range(dim[0]):
        for j in range(dim[1]):
            for k in range(dim[2]):
                #edges[i,j,k] = (edges_x[i,j,k] and edges_y[i,j,k]) or (edges_x[i,j,k] and edges_z[i,j,k]) or (edges_y[i,j,k] and edges_z[i,j,k])
                edges[i,j,k] = (edges_x[i,j,k]) or (edges_y[i,j,k]) or (edges_z[i,j,k])
    
    
    return edges

def accumulate_gradients(r_table, grayImage):
    '''
    Perform a General Hough Transform with the given image and R-table
    '''
    
    #Get edges matrix from Canny 
    edges = canny_edges_3d(grayImage) 
    
    #Get gradient angles
    phi, psi = gradient_orientation(edges)
    
    accumulator = np.zeros(grayImage.shape)
    accum_i = 0
    accum_j = 0 
    accum_k = 0

    #print (datetime.datetime.now())

    for (i,j,k),value in np.ndenumerate(edges):
        if value:
            for r in r_table[(int(phi[i,j,k]), int(psi[i,j,k]))]:
                accum_i, accum_j, accum_k = i+r[0], j+r[1], k+r[2]
                if accum_i < accumulator.shape[0] and accum_j < accumulator.shape[1] and accum_k < accumulator.shape[2]:
                    accumulator[int(accum_i), int(accum_j), int(accum_k)] += 1
                
    #print (datetime.datetime.now())

    return accumulator

def general_hough_closure(reference_image):
    '''
    Generator function to create a closure with the reference image and origin
    at the center of the reference image
    
    Returns a function f, which takes a query image and returns the accumulator
    '''
    
    referencePoint = (reference_image.shape[0]/2, reference_image.shape[1]/2, reference_image.shape[2]/2)
    
    #print("Reference Point: ", referencePoint)
    
    r_table = build_r_table(reference_image, referencePoint)
    
    def f(query_image):
        return accumulate_gradients(r_table, query_image)
        
    return f

def n_max(a, n):
    '''
    Return the N max elements and indices in a
    '''
    indices = (-a.ravel()).argsort()[:n]
    indices = (np.unravel_index(i, a.shape) for i in indices)
    return [(a[i], i) for i in indices]
 
def test_general_hough(gh, reference_image, query):
    '''
    Uses a GH closure to detect shapes in an image and create nice output
    '''
    #query_image = imread(query, flatten=True)
    query_image = query
    query_dim = np.shape(query)
    reference_dim = np.shape(reference_image)

    accumulator = gh(query_image)

    return accumulator


#===================================================================================================
#****************************************** START OF GHT *******************************************
#===================================================================================================
def GHT(ac_num):
#===================================================================================================
#Obtaining the list of references to use and bounding the Region of Interest
#===================================================================================================
    #Get the references that will be used for Cervical Spine Vertebrae Detection
    reference_acs = []

    for file_name in os.listdir("no_fractures"):
        #The name should have "reference" in it but no "edge" in it
        if file_name.find("reference.pkl") != -1 and file_name.find("edge") == -1:
            #The reference list should not include the reference from the current ac_num
            if file_name.find(str(ac_num)) == -1:
                reference_acs.append(file_name)
    
    print("Accession Number: ", ac_num)
    
    
    #Open Downsized Pickle File Containing DICOM Scan
    dicom_dwn4x_pp = cPickle.load(open("no_fractures/dicom_3d_" + ac_num + "_dwn4x.pkl","rb"),encoding = 'latin1')
    dicom_dwn4x_pp_dim = np.shape(dicom_dwn4x_pp)
    print("Size of Downsized Dicom Input: ", dicom_dwn4x_pp_dim)

    #Specify Region of Interest
    x1 = 5
    x2 = 65
    y1 = 15
    y2 = 80

    
    #Get specific region of focus (based on prior information)
    dicom_dwn4x = dicom_dwn4x_pp[x1:x2,y1:y2,:] #dicom_dwn4x contains the specific region of focus
    dicom_dwn4x_dim = np.shape(dicom_dwn4x)
    print("Size of Relevant Dicom (Prior Info): ", dicom_dwn4x_dim)
   

#===================================================================================================
#Obtain Final Accumulator Matrix through Max-Pooling of Individual Accumulator Matrices
#===================================================================================================
    #Initialize Accumulator that will be used to get top points
    accumulator = np.zeros(dicom_dwn4x_dim)
    
    for reference_ac in reference_acs:
        print("Current Reference: ", reference_ac)
        
        #Open up the Reference that is used as the reference image
        reference = cPickle.load(open("no_fractures/" + reference_ac,"rb"),encoding = 'latin1')
        #print("Size of Reference Image: ", np.shape(reference))

        detect_s = general_hough_closure(reference)
        
        #Use max pooling on accumulator matrix
        temp_accumulator = test_general_hough(detect_s, reference, dicom_dwn4x)
        
        accumulator = np.maximum(accumulator,temp_accumulator)
    
    final_accumulator = accumulator
    
    
#===================================================================================================
#Blurring the Accumulator Matrix and Query Edge Image
#===================================================================================================
    #Blur the final accumulator matrix
    global std_dev
    std_dev = 1.0
    final_accumulator = gaussian_filter(final_accumulator,sigma = std_dev, order = 0)

    #Blur the edge image for the whole dwn4x
    global std_dev_edges
    std_dev_edges = 0.5
    
    query_edges = canny_edges_3d(dicom_dwn4x_pp)

    query_edges_dim = np.shape(query_edges)
    
    query_edges_blurred = gaussian_filter(np.multiply(query_edges,50),sigma = std_dev_edges, order = 0)

    #Sanity Check regarding number of negative edges
    neg = 0
    for i in range(query_edges_dim[0]):
        for j in range(query_edges_dim[1]):
            for k in range(query_edges_dim[2]):
                if query_edges_blurred[i,j,k] < 0:
                    neg = neg + 1
                    
    print("Number of Negative Edges: ", neg)


#===================================================================================================
#Initial Plots and Top 40 Points for Visualization Purposes
#===================================================================================================
    #Plot up to top 40 points
    fig = plt.figure(num = ac_num + "_accumulator_sigma" + str(std_dev) + "_edge_sigma_" + str(std_dev_edges) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD), figsize = (24,12))
    plt.gray()

    fig.suptitle(ac_num + "_accumulator_sigma" + str(std_dev) + "_edge_sigma_" + str(std_dev_edges) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD))

    fig.add_subplot(2,4,1)
    plt.title('Query Image')
    plt.imshow(dicom_dwn4x_pp[:,:,dicom_dwn4x_pp_dim[2]//2])
    
    fig.add_subplot(2,4,2)
    plt.title('Query Image Edges')
    plt.imshow(query_edges[:,:,dicom_dwn4x_pp_dim[2]//2])
    
    fig.add_subplot(2,4,3)
    plt.title('Query Image Edges Blurred')
    plt.imshow(query_edges_blurred[:,:,dicom_dwn4x_pp_dim[2]//2])
    
    fig.add_subplot(2,4,4)
    plt.title('Final Accumulator')
    plt.imshow(final_accumulator[:,:,dicom_dwn4x_dim[2]//2])
     
    fig.add_subplot(2,4,5)
    plt.title('Detection of Top 40 Points')
    plt.imshow(dicom_dwn4x_pp[:,:,dicom_dwn4x_dim[2]//2])


    #Get top 40 results that can be filtered out
    m = n_max(final_accumulator, 40)

    points = []
    x_pts = [] 
    y_pts = []
    z_pts = []
    
    for pt in m:
        points.append((pt[1][0] + x1,pt[1][1] + y1,pt[1][2], int(pt[0])))
    
        x_pts.append(pt[1][0]+x1)
        y_pts.append(pt[1][1]+y1) 
        z_pts.append(pt[1][2])
    
    plt.scatter(y_pts,x_pts, marker='.', color='r')
        
    #print ("Top 40 Most Likely Points (x,y,z,certainty): ", points)


#===================================================================================================
#Non-maximal suppression
#===================================================================================================
    #Plot NMS points
    fig.add_subplot(2,4,6)
    plt.title('Non-Maximal Suppression and Optimal Points')
    plt.imshow(dicom_dwn4x_pp[:,:,dicom_dwn4x_pp_dim[2]//2])

    #Perform non-maximal suppression
    nms_pts = []
    
    for pt in points:
        if len(nms_pts) == 0:
            nms_pts.append(pt)
        else:
            counter = 0
            for i in range(len(nms_pts)):
                if math.sqrt((nms_pts[i][0]-pt[0])**2 + (nms_pts[i][1]-pt[1])**2 + (nms_pts[i][2]-pt[2])**2) > 10:
                    counter = counter + 1
                else:
                    if pt[3] > nms_pts[i][3]:
                        nms_pts[i] = pt
            
            if counter == len(nms_pts):
                nms_pts.append(pt)
        
    print("Non-Maximal Suppression Points: ", nms_pts)
  

#===================================================================================================
#Normalized Cross Correlation and Heat Map Generation
#===================================================================================================
    #Sliding reference across volume around detected points to find accurate point
    optimal_pt = [0,0]
    max_cross_correl_val = -float('inf')
    
    heat_maps = []
    
    #Generate the edge references if they do not exist in the directory "no_Fractures"
    for reference_ac in reference_acs:
        if not os.path.isfile("no_fractures/edge_" + str(std_dev_edges) + "_" + reference_ac):
            reference_vol_pp1 = cPickle.load(open("no_fractures/" + reference_ac,"rb"),encoding = 'latin1')
            reference_vol_pp2 = np.array(reference_vol_pp1)
            
            reference_vol_edges = canny_edges_3d(reference_vol_pp2)
            reference_vol_edges_blurred = gaussian_filter(np.multiply(reference_vol_edges,50),sigma = std_dev_edges, order = 0)
            
            cPickle.dump(reference_vol_edges_blurred, open("no_fractures/edge_" + str(std_dev_edges) + "_" + reference_ac,"wb"))
            #print("no_fractures/edge_" + reference_ac)
            
    
    for pt in nms_pts:
        heat_map = np.zeros((9,9,3))
        print("The point being investigated is: ", pt)

        for i in range(-4,5):
            for j in range(-4,5):
                for k in range(-1,2):
                    cross_correl_val = 0
                    
                    for reference_ac in reference_acs:
                        reference_vol_pp = cPickle.load(open("no_fractures/edge_" + str(std_dev_edges) + "_" + reference_ac,"rb"),encoding = 'latin1')
                        #reference_dim is the dimension of the edge reference
                        reference_dim = np.shape(reference_vol_pp)
                        reference_vol = np.ndarray.flatten(reference_vol_pp)
                        
                        #Get bounds to compare on the query image
                        x1 = pt[0] - reference_dim[0]//2 + i
                        x2 = x1 + reference_dim[0]
                        
                        y1 = pt[1] - reference_dim[1]//2 + j
                        y2 = y1 + reference_dim[1]
                        
                        z1 = pt[2] - reference_dim[2]//2 + k
                        z2 = z1 + reference_dim[2]
                        
                        #Use the Canny edge version of the query image for cross correlation
                        query_vol_pp = np.array(query_edges_blurred[x1:x2,y1:y2,z1:z2])
                        query_vol = np.ndarray.flatten(query_vol_pp)
 
                        query_dim = np.shape(query_vol_pp)
                        
                        #Exit current slide location if out of bounds
                        if x1 < 0 or y1 < 0 or z1 < 0:
                            break


                        #Use norms to normalize the vectors for cross-correlation
                        #print(np.linalg.norm(reference_vol))
                        #print(np.linalg.norm(query_vol))
                        reference_vol_norm = reference_vol/np.linalg.norm(reference_vol)
                        query_vol_norm = query_vol/np.linalg.norm(query_vol)
                        
                        
                        if (np.dot(reference_vol_norm, query_vol_norm)) < 0:
                            print("ALERT ALERT ALERT")
                        
                        cross_correl_val = cross_correl_val + np.dot(reference_vol_norm, query_vol_norm)
                    
                    heat_map[i+4,j+4,k+1] = cross_correl_val
                    if cross_correl_val > max_cross_correl_val:
                        #print(max_cross_correl_val)
                        max_cross_correl_val = cross_correl_val
                        #print("The cross correlation value is: ", cross_correl_val)
                        optimal_pt = [pt[0]+i,pt[1]+j, pt[2]+k]
                        #print("The optimal point currently is: ", optimal_pt)
        
        
        #Append heat_map
        heat_maps.append(heat_map)
                    
    print("The Final Detection point is: ",optimal_pt)

  
    #Plot non-maximal suppression points
    nms_x_pts = [] 
    nms_y_pts = []
    nms_z_pts = []
    
    for pt in nms_pts:
        nms_x_pts.append(pt[0])
        nms_y_pts.append(pt[1]) 
        nms_z_pts.append(pt[2])    


    plt.scatter(nms_y_pts,nms_x_pts, marker='o', color='g')
    plt.scatter(optimal_pt[1],optimal_pt[0], marker='X', color='m')
    
    
    #Add plot for heat map
    for i in range(2):
        try:
            heat_map = heat_maps[i]
            heat_map_norm = heat_map
            fig.add_subplot(2,4,7+i)
            plt.title('Heat Map')
            plt.imshow(heat_map_norm[:,:,1])
        except:
            pass
    
    #plt.show()
    
    #Save Figure
    #plt.savefig(ac_num + "_sigma" + str(std_dev) + ".png")
    
    return optimal_pt

#===================================================================================================
#******************************************* END OF GHT ********************************************
#===================================================================================================


#===================================================================================================
#===================================================================================================
if __name__ == '__main__':
    os.chdir("C:\\Users\\yoons\\Documents\\4th Year Semester 2\\ESC499 - Thesis\\Undergraduate_Thesis_Scripts\\DicomSubsampling")
    

#===================================================================================================
#Process the accession numbers that are present and put it into a list
#===================================================================================================
    ac_nums_pp = os.listdir("no_fractures/")
    ac_nums = []
    
    for ac_num_pp in ac_nums_pp:
        str1 = ac_num_pp.split("dicom_3d_")[1]
        str2 = str1.split("_")[0]
        
        ac_nums.append(str2)
    
    #print(ac_nums)
    #print(len(ac_nums))
#===================================================================================================
#Read in ground truth values from the ground_truth_detection_pts.xlsx spreadsheet
#===================================================================================================
    #Get the detection results for the validation set
    book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts.xlsx")
    sheet = book.active
    row_count = sheet.max_row
    
    ground_truth = {}

    for i in range(3,row_count): #divided by 3 as a test 
        
        ac_num_loc = sheet.cell(row = i,column = 1)
        ac_num = str(ac_num_loc.value)
        
        x = sheet.cell(row = i, column = 2).value
        y = sheet.cell(row = i, column = 3).value
        z = sheet.cell(row = i, column = 4).value
        
        if (x != None) and (y != None) and (z != None):
            ground_truth[ac_num] = [x,y,z]


#===================================================================================================
#Compute Detection Points, compare with ground truth to get error and detection rate
#===================================================================================================
    error = 0
    correct_detections = 0
    total_detections = 0
    incorrect_ac_num = []
    detection_pt_info = {}
    
    #Go through GHT for the validation set
    for ac_num in ac_nums:
        if ac_num in ground_truth.keys():
            total_detections = total_detections + 1
            
            optimal_pt = GHT(ac_num)
            print("Detected Optimal Point: ", optimal_pt)
            print("Ground Truth Point: ", ground_truth[ac_num])
        
            curr_error = abs(np.linalg.norm(np.subtract(optimal_pt,ground_truth[ac_num])))**2 
            error = error + curr_error
            
            #Can adjust threshold for correct detection accordingly
            if curr_error <= 12:
                correct_detections = correct_detections + 1
            else:
                incorrect_ac_num.append(ac_num)
                
            #Keep record of the information
            detection_pt_info[ac_num] = optimal_pt	
    
    plt.show()
    
    print("======================================")
    print("********SUMMARY OF PERFORMANCE********")
    print("======================================")

    print("Min Canny Threshold: ", MIN_CANNY_THRESHOLD)
    print("Max Canny Threshold: ", MAX_CANNY_THRESHOLD)
    print("Sigma Accumulator: ", std_dev)
    print("Sigma Edges: ", std_dev_edges)

    print("The squared error for this trial on the validation set is :", error)
    print("The detection rate is: " + str(correct_detections) + "/" + str(total_detections))
    
    print("The Access Numbers for Incorrect Detections are: ", incorrect_ac_num)
    print("Detection Point Information: ", detection_pt_info)
#===================================================================================================
#===================================================================================================