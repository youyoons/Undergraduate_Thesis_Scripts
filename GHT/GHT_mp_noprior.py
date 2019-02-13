import os
import numpy as np
import math
import matplotlib.pyplot as plt
from collections import defaultdict
from scipy.misc import imread
from skimage.feature import canny
from scipy.ndimage.filters import sobel
from mpl_toolkits.mplot3d import Axes3D
import cv2 as cv
from scipy import signal
from scipy.ndimage.filters import gaussian_filter
import datetime
import openpyxl
import random
import cython
from multiprocessing import Pool

try:
    import cPickle
except ImportError:
    import pickle as cPickle


def gradient_orientation(image):
    '''
    Calculate the gradient orientation for edge point in the image
    '''
    #scipy.ndimage.sobel
    dx = sobel(image, axis=0, mode='constant')	
    dy = sobel(image, axis=1, mode='constant')
    dz = sobel(image, axis=1, mode='constant')
    
    #For 3D instead of a single gradient value, we need two angles that define a normal vector
    #Phi is the angle between the positive x-axis to the projection of the normal vector the x-y plane (around +z)
    #Psi is the angle between the positive z-axis to the normal vector
    
    phi = np.arctan2(dy ,dx) * 180 / np.pi
    psi = np.arctan2(np.sqrt(dx*dx + dy*dy), dz) * 180 / np.pi
    

    gradient = np.zeros(image.shape)
    
    return phi, psi

def build_r_table(image, origin):
    '''
    Build the R-table from the given shape image and a reference point
    '''
    edges = canny_edges_3d(image)
    
    #Takes (47,40) Edges and calculates the gradients using sobel
    phi, psi = gradient_orientation(edges)
    #print("Phi Dim: ", phi.shape)
    
    r_table = defaultdict(list)
    for (i,j,k),value in np.ndenumerate(edges):
        if value:
            r_table[(int(phi[i,j,k]),int(psi[i,j,k]))].append((origin[0]-i, origin[1]-j, origin[2] - k))
    
    
    return r_table

def canny_edges_3d(grayImage):
    dim = np.shape(grayImage)
    
    edges_x = np.zeros(grayImage.shape, dtype=bool) 
    edges_y = np.zeros(grayImage.shape, dtype=bool) 
    edges_z = np.zeros(grayImage.shape, dtype=bool) 
    edges = np.zeros(grayImage.shape, dtype=bool) 
    

    for i in range(dim[0]):
        edges_x[i,:,:] = canny(grayImage[i,:,:], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = std_dev_canny)
   
    for j in range(dim[1]):
        edges_y[:,j,:] = canny(grayImage[:,j,:], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = std_dev_canny)
        
    for k in range(dim[2]):
        edges_z[:,:,k] = canny(grayImage[:,:,k], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = std_dev_canny)
    
    
    for i in range(dim[0]):
        for j in range(dim[1]):
            for k in range(dim[2]):
                edges[i,j,k] = (edges_x[i,j,k]) or (edges_y[i,j,k]) or (edges_z[i,j,k])
    
    
    return edges

#grayImage is queryImage
def accumulate_gradients(r_table, grayImage):
    '''
    Perform a General Hough Transform with the given image and R-table
    '''
    
    #Get edges matrix from Canny 
    edges = canny_edges_3d(grayImage) 
    
    #Get gradient angles
    phi, psi = gradient_orientation(edges)
    
    accumulator = np.zeros(grayImage.shape)

    accum_i = 0
    accum_j = 0 
    accum_k = 0

    edges_dim = np.shape(edges)

    print(datetime.datetime.now())
    
    
    for (i,j,k),value in np.ndenumerate(edges):
        if value: 
            for r in r_table[(int(phi[i,j,k]), int(psi[i,j,k]))]:
                #iterations = iterations + 1
                accum_i, accum_j, accum_k = i+r[0], j+r[1], k+r[2]
                if accum_i < accumulator.shape[0] and accum_j < accumulator.shape[1] and accum_k < accumulator.shape[2]:
                    accumulator[int(accum_i), int(accum_j), int(accum_k)] += 1 
  
 
    print(datetime.datetime.now())  
    

    return accumulator

def general_hough_closure(reference_image):
    '''
    Generator function to create a closure with the reference image and origin
    at the center of the reference image
    
    Returns a function f, which takes a query image and returns the accumulator
    '''
    
    referencePoint = (reference_image.shape[0]/2, reference_image.shape[1]/2, reference_image.shape[2]/2)
    
    r_table = build_r_table(reference_image, referencePoint)
    
    def f(query_image):
        return accumulate_gradients(r_table, query_image)
        
    return f

def n_max(a, n):
    '''
    Return the N max elements and indices in a
    '''
    indices = (-a.ravel()).argsort()[:n]
    indices = (np.unravel_index(i, a.shape) for i in indices)
    return [(a[i], i) for i in indices]
 
def test_general_hough(gh, reference_image, query):
    '''
    Uses a GH closure to detect shapes in an image and create nice output
    '''
    query_image = query
    query_dim = np.shape(query)
    reference_dim = np.shape(reference_image)

    accumulator = gh(query_image)

    return accumulator


#===================================================================================================
#****************************************** START OF GHT *******************************************
#===================================================================================================
def GHT(ac_num):
#===================================================================================================
#Obtaining the list of references to use and bounding the Region of Interest
#===================================================================================================
    #Get the references that will be used for Cervical Spine Vertebrae Detection
    reference_acs = []

    image_file_name = ac_num + "_accumulator_sigma_" + str(std_dev) + "_edge_sigma_" + str(std_dev_edges)  + "_canny_sigma_" + str(std_dev_canny) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD)

    for file_name in os.listdir("no_fractures"):
        #The name should have "reference" in it but no "edge" in it
        if file_name.find("reference.pkl") != -1 and file_name.find("edge") == -1:
            #The reference list should not include the reference from the current ac_num
            if file_name.find(str(ac_num)) == -1:
                reference_acs.append(file_name)
    
    print("Accession Number: ", ac_num)
    
    
    #Open Downsized Pickle File Containing DICOM Scan
    try:
        dicom_dwn4x_pp = cPickle.load(open("no_fractures/dicom_3d_" + ac_num + "_dwn4x.pkl","rb"),encoding = 'latin1')
    except:
        dicom_dwn4x_pp = cPickle.load(open("no_fractures/dicom_3d_" + ac_num + "_dwn4x.pkl","rb"))
    dicom_dwn4x_pp_dim = np.shape(dicom_dwn4x_pp)
    print("Size of Downsized Dicom Input: ", dicom_dwn4x_pp_dim)

#**************************************************************************************************************************
    #Specify Region of Interest (Hard Blocking of Region Based on Prior Information)
    x1 = 0
    x2 = 58
    y1 = 17
    y2 = 85
#**************************************************************************************************************************

    
    #Get specific region of focus (based on prior information)
    dicom_dwn4x = dicom_dwn4x_pp[x1:x2,y1:y2,:] #dicom_dwn4x contains the specific region of focus
    dicom_dwn4x_dim = np.shape(dicom_dwn4x)
    print("Size of Relevant Dicom (Prior Info): ", dicom_dwn4x_dim)
   

#===================================================================================================
#Obtain Final Accumulator Matrix through Max-Pooling of Individual Accumulator Matrices
#===================================================================================================
    #Initialize Accumulator that will be used to get top points
    accumulator = np.zeros(dicom_dwn4x_dim)
    
    
    #Choose N number of references
    #random_reference_acs = []
    random_reference_acs = reference_acs[:]

    
    #while len(random_reference_acs) < 5: 
    #    index = random.randint(0,len(reference_acs)-1)
        
    #    if reference_acs[index] not in random_reference_acs:
    #        random_reference_acs.append(reference_acs[index])
    
    for reference_ac in random_reference_acs:
        print("Current Reference: ", reference_ac)
        
        #Open up the Reference that is used as the reference image
        try:
            reference = cPickle.load(open("no_fractures/" + reference_ac,"rb"),encoding = 'latin1')
        except:
            reference = cPickle.load(open("no_fractures/" + reference_ac,"rb"))
        #print("Size of Reference Image: ", np.shape(reference))

        detect_s = general_hough_closure(reference)
        
        #Use max pooling on accumulator matrix
        temp_accumulator = test_general_hough(detect_s, reference, dicom_dwn4x)
        
        accumulator = np.maximum(accumulator,temp_accumulator)
    
    final_accumulator = accumulator


#===================================================================================================
#Blurring the Accumulator Matrix and Query Edge Image
#===================================================================================================
    #Blur the final accumulator matrix
    final_accumulator = gaussian_filter(final_accumulator,sigma = std_dev, order = 0)

    #Blur the edge image for the whole dwn4x
    query_edges = canny_edges_3d(dicom_dwn4x_pp)

    query_edges_dim = np.shape(query_edges)
    
    query_edges_blurred = gaussian_filter(np.multiply(query_edges,50),sigma = std_dev_edges, order = 0)


#===================================================================================================
#Initial Plots and Top 40 Points for Visualization Purposes
#===================================================================================================
    plot_z = ground_truth[ac_num][2]

    #Plot up to top 40 points
    fig = plt.figure(num = image_file_name, figsize = (24,12))
    plt.gray()

    fig.suptitle(image_file_name)

    fig.add_subplot(2,4,1)
    plt.title('Query Image [Slice: ' + str(plot_z) + ']')
    #plt.imshow(dicom_dwn4x_pp[:,:,dicom_dwn4x_pp_dim[2]//2])
    plt.imshow(dicom_dwn4x_pp[:,:,plot_z])
    
    fig.add_subplot(2,4,2)
    plt.title('Query Image Edges')
    #plt.imshow(query_edges[:,:,dicom_dwn4x_pp_dim[2]//2])
    plt.imshow(query_edges[:,:,plot_z])
    
    fig.add_subplot(2,4,3)
    plt.title('Query Image Edges Blurred')
    #plt.imshow(query_edges_blurred[:,:,dicom_dwn4x_pp_dim[2]//2])
    plt.imshow(query_edges_blurred[:,:,plot_z])
    
    fig.add_subplot(2,4,4)
    plt.title('Final Accumulator')
    #plt.imshow(final_accumulator[:,:,dicom_dwn4x_dim[2]//2])
    plt.imshow(final_accumulator[:,:,plot_z])
     
    fig.add_subplot(2,4,5)
    plt.title('Detection of Top 40 Points')
    #plt.imshow(dicom_dwn4x_pp[:,:,dicom_dwn4x_dim[2]//2])
    plt.imshow(dicom_dwn4x_pp[:,:,plot_z])


    #Get top 40 results that can be filtered out
    m = n_max(final_accumulator, 40)

    points = []
    x_pts = [] 
    y_pts = []
    z_pts = []
    
    for pt in m:
        points.append((pt[1][0] + x1,pt[1][1] + y1,pt[1][2], int(pt[0])))
    
        x_pts.append(pt[1][0]+x1)
        y_pts.append(pt[1][1]+y1) 
        z_pts.append(pt[1][2])
    
    plt.scatter(y_pts,x_pts, marker='.', color='r')
     
    
    #Take the top K average 
    k = 5
    k_sum_pp = np.zeros(3)
    for index in range(k):
        k_sum_pp = np.add(k_sum_pp, m[index][1])
        print(m[index])
    
    optimal_pt = (int(k_sum_pp[0]//k) + x1,int(k_sum_pp[1]//k) + y1,int(k_sum_pp[2]//k))


#===================================================================================================
#Non-maximal suppression
#===================================================================================================
    #Plot NMS points
    fig.add_subplot(2,4,6)
    plt.title('Non-Maximal Suppression and Optimal Points')
    #plt.imshow(dicom_dwn4x_pp[:,:,dicom_dwn4x_pp_dim[2]//2])
    plt.imshow(dicom_dwn4x_pp[:,:,plot_z])

    #Perform non-maximal suppression
    nms_pts = []
    
    for pt in points:
        if len(nms_pts) == 0:
            nms_pts.append(pt)
        else:
            counter = 0
            for i in range(len(nms_pts)):
                if math.sqrt((nms_pts[i][0]-pt[0])**2 + (nms_pts[i][1]-pt[1])**2 + (nms_pts[i][2]-pt[2])**2) > 10:
                    counter = counter + 1
                else:
                    if pt[3] > nms_pts[i][3]:
                        nms_pts[i] = pt
            
            if counter == len(nms_pts):
                nms_pts.append(pt)
        
    print("Non-Maximal Suppression Points: ", nms_pts)
  

#===================================================================================================
#Normalized Cross Correlation and Heat Map Generation
#===================================================================================================
    #Sliding reference across volume around detected points to find accurate point
    #optimal_pt = [0,0]
    max_cross_correl_val = -float('inf')
    

    print("The Final Detection point is: ",optimal_pt)

  
    #Plot non-maximal suppression points
    nms_x_pts = [] 
    nms_y_pts = []
    nms_z_pts = []
    
    for pt in nms_pts:
        nms_x_pts.append(pt[0])
        nms_y_pts.append(pt[1]) 
        nms_z_pts.append(pt[2])    


    plt.scatter(nms_y_pts,nms_x_pts, marker='o', color='g')
    
    plt.scatter(optimal_pt[1],optimal_pt[0], marker='X', color='m')
    
    #Put on ground truth point on NMS + Optimal Point Plot
    plt.scatter(ground_truth[ac_num][1], ground_truth[ac_num][0], marker= 'o', color = 'c')
    
    
    
    #Save Figure
    print(image_dir_name)
    plt.savefig(image_dir_name + "/" + image_file_name + ".png")
    
    return optimal_pt

#===================================================================================================
#******************************************* END OF GHT ********************************************
#===================================================================================================


#===================================================================================================
#===================================================================================================
if __name__ == '__main__':
    #os.chdir("C:\\Users\\yoons\\Documents\\ESC499\\Undergraduate_Thesis_Scripts\\DicomSubsampling")
    os.chdir("../DicomSubsampling")    

    plt.close()
#===================================================================================================
#Process the accession numbers that are present and put it into a list
#===================================================================================================
    ac_nums_pp = os.listdir("no_fractures/")
    ac_nums = []
    
    for ac_num_pp in ac_nums_pp:
        if "reference" not in ac_num_pp:
            str1 = ac_num_pp.split("dicom_3d_")[1]
            str2 = str1.split("_")[0]
            
            ac_nums.append(str2)

#===================================================================================================
#Read in ground truth values from the ground_truth_detection_pts.xlsx spreadsheet
#===================================================================================================
    #Get the detection results for the validation set
    book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts_all.xlsx")
    sheet = book.active
    row_count = sheet.max_row
    
    global ground_truth
    ground_truth = {}

    for i in range(3,row_count+1):
        
        ac_num_loc = sheet.cell(row = i,column = 1)
        ac_num = str(ac_num_loc.value)
        
        x = sheet.cell(row = i, column = 2).value
        y = sheet.cell(row = i, column = 3).value
        z = sheet.cell(row = i, column = 4).value
        
        if (x != None) and (y != None) and (z != None):
            ground_truth[ac_num] = [x,y,z]
    
#===================================================================================================
#Compute Detection Points, compare with ground truth to get error and detection rate
#===================================================================================================
    global std_dev
    global std_dev_edges
    global MIN_CANNY_THRESHOLD
    global MAX_CANNY_THRESHOLD
    global std_dev_canny
    global image_file_name
    global image_dir_name
    
    #Set Hyperparameters to be validated with validation set
    std_devs = [1.0]
    std_devs_edges = [0]
    min_cannys = [30,40,50,60]
    max_cannys = [140,160,180,200]


    std_dev_canny = 0.5
    
    for std_dev in std_devs:
        for std_dev_edges in std_devs_edges:
            for MIN_CANNY_THRESHOLD in min_cannys:
                for MAX_CANNY_THRESHOLD in max_cannys:
                    error = 0
                    correct_detections = 0
                    incorrect_ac_num = []
                    detection_pt_info = {}
                    
                    image_dir_name = "accumulator_sigma_" + str(std_dev) + "_edge_sigma_" + str(std_dev_edges) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD)
                    
                    print("Currently on: " + image_dir_name)
                    
                    if os.path.isdir(image_dir_name) != 1:
                        os.mkdir(image_dir_name)
                    else:
                        continue
                    

                    #Get the ac_num to put into multi processing
                    multi_proc_ac_num = []
                    
                    for ac_num in ac_nums:
                        if ac_num in ground_truth.keys():
                            multi_proc_ac_num.append(ac_num)
                    
                    print(multi_proc_ac_num)
                    
                    #Get optimal points through multi processing
                    p = Pool(processes = 25)
               
                    optimal_pts = p.map(GHT,multi_proc_ac_num)
                    
                    optimal_pts_dict = {}
                    #Put into dictionary
                    for i in range(len(multi_proc_ac_num)):
                        optimal_pts_dict[multi_proc_ac_num[i]] = optimal_pts[i]
                    
                    print(optimal_pts_dict)

                    #Go through GHT for the validation set
                    for ac_num in ac_nums:
                        if ac_num in ground_truth.keys():
                            
                            optimal_pt = optimal_pts_dict[ac_num]
                            print("Detected Optimal Point: ", optimal_pt)
                            print("Ground Truth Point: ", ground_truth[ac_num])
                        
                            curr_error = abs(np.linalg.norm(np.subtract(optimal_pt,ground_truth[ac_num])))**2 
                            error = error + curr_error
                            
                            #Can adjust threshold for correct detection accordingly
                            if curr_error <= 20.0:
                                correct_detections = correct_detections + 1
                            else:
                                incorrect_ac_num.append(ac_num)
                                
                            #Keep record of the information
                            detection_pt_info[ac_num] = [optimal_pt, ground_truth[ac_num], curr_error]
                    
                    plt.close()
                    
                
                    '''
                    print("======================================")
                    print("********SUMMARY OF PERFORMANCE********")
                    print("======================================")
                
                    print("Min Canny Threshold: ", MIN_CANNY_THRESHOLD)
                    print("Max Canny Threshold: ", MAX_CANNY_THRESHOLD)
                    print("Sigma Canny: ", std_dev_canny)
                    print("Sigma Accumulator: ", std_dev)
                    print("Sigma Edges: ", std_dev_edges)
                
                    print("The squared error for this trial on the validation set is :", error)
                    print("The detection rate is: " + str(correct_detections) + "/" + str(len(ground_truth.keys())))
                    
                    print("The Accession Numbers for Incorrect Detections are: ", incorrect_ac_num)
                    print("Detection Point Information: ", detection_pt_info)
                    '''

                    #Output General Information to File
                    f = open(image_dir_name + "/summary.txt","w")
                    f.write("======================================\n")
                    f.write("********SUMMARY OF PERFORMANCE********\n")
                    f.write("======================================\n")
                    f.write("Min Canny Threshold: %s \n" % str(MIN_CANNY_THRESHOLD))
                    f.write("Max Canny Threshold: %s \n" % str(MAX_CANNY_THRESHOLD))
                    f.write("Sigma Canny: %s \n" % str(std_dev_canny))
                    f.write("Sigma Accumulator: %s \n" % str(std_dev))
                    f.write("Sigma Edges: %s \n\n" % str(std_dev_edges))
                    
                    f.write("The squared error for this trial on the validation set is: %s \n\n" % str(error))
                    
                    f.write("The number of correct detections is %s" % correct_detections)
                    f.write("/%s \n\n" % str(len(ground_truth.keys())))
                    
                    f.write("Incorrect Accession Numbers: \n")
                    for item in incorrect_ac_num:
                        f.write("%s  " % str(item))

                    f.write("\n\n")
                    f.write("Below are the detections points: \n")
                    for key in detection_pt_info.keys():
                        f.write("AC Num: %s    Detected Point: " % str(key))
                        
                        info = detection_pt_info[key]
                        f.write("%s    Actual Point: " % str(info[0]))
                        f.write("%s    Error: " %str(info[1]))
                        f.write("%s" % info[2])
                        f.write("\n")
                    
                    f.close()
                    
                    #Create Excel Spreadsheet with Detection Information
                    wb = openpyxl.Workbook()
                    dest_filename = '../GHT/detection_pts_trial_' + str(MIN_CANNY_THRESHOLD) + '_' + str(MAX_CANNY_THRESHOLD) + '_' + str(std_dev_canny) + '_' + str(std_dev) + '_' + str(std_dev_edges) + '.xlsx'


                    ws1 = wb.active
                    ws1.append(['ac_num','x','y','z'])
                    for key in detection_pt_info.keys():
                        detected_pt = detection_pt_info[key][0]
                        row = [key,detected_pt[0],detected_pt[1],detected_pt[2]]

                        ws1.append(row)
                    wb.save(dest_filename)
                
                    
#===================================================================================================
#===================================================================================================
