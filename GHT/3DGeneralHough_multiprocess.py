import os
import numpy as np
import math
import matplotlib.pyplot as plt
from collections import defaultdict
from scipy.misc import imread
from skimage.feature import canny
from scipy.ndimage.filters import sobel
from mpl_toolkits.mplot3d import Axes3D
import cv2 as cv
from scipy import signal
from scipy.ndimage.filters import gaussian_filter
import datetime
import openpyxl
import random
import cython
from multiprocessing import Pool
import time
import functions
try:
    import cPickle
except ImportError:
    import pickle as cPickle



#===================================================================================================
#===================================================================================================
if __name__ == '__main__':
    os.chdir("c:\\users\\yoonsun you\\documents\\thesis_semester2\\undergraduate_thesis_scripts\\DicomSubsampling")
    
    plt.close()
#===================================================================================================
#Process the accession numbers that are present and put it into a list
#===================================================================================================
    ac_nums_pp = os.listdir("no_fractures/")
    ac_nums = []
    
    for ac_num_pp in ac_nums_pp:
        if "reference" not in ac_num_pp:
            str1 = ac_num_pp.split("dicom_3d_")[1]
            str2 = str1.split("_")[0]
            
            ac_nums.append(str2)

#===================================================================================================
#Read in ground truth values from the ground_truth_detection_pts.xlsx spreadsheet
#===================================================================================================
    #Get the detection results for the validation set
    book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts_all.xlsx")
    sheet = book.active
    row_count = sheet.max_row
    
    ground_truth = {}

    for i in range(3,row_count): #divided by 3 as a test 
        
        ac_num_loc = sheet.cell(row = i,column = 1)
        ac_num = str(ac_num_loc.value)
        
        x = sheet.cell(row = i, column = 2).value
        y = sheet.cell(row = i, column = 3).value
        z = sheet.cell(row = i, column = 4).value
        
        if (x != None) and (y != None) and (z != None):
            ground_truth[ac_num] = [x,y,z]
    
#===================================================================================================
#Compute Detection Points, compare with ground truth to get error and detection rate
#===================================================================================================
    global std_dev
    global std_dev_edges
    global MIN_CANNY_THRESHOLD
    global MAX_CANNY_THRESHOLD
    global std_dev_canny
    global image_file_name
    global image_dir_name
    
    #Set Hyperparameters to be validated with validation set
    std_devs = [1.0,1.5,2.0]
    #std_devs_edges = [0,0.5,1.0,1.5,2.0]
    std_devs_edges = [0]
    #min_cannys = [20,30,40,50,60]
    min_cannys = [30,40,50,60]
    #max_cannys = [160,180,200,220,240,260]
    max_cannys = [160,180,200,220]

    std_dev_canny = 0.5
    
    for std_dev in std_devs:
        for std_dev_edges in std_devs_edges:
            for MIN_CANNY_THRESHOLD in min_cannys:
                for MAX_CANNY_THRESHOLD in max_cannys:
                    error = 0
                    correct_detections = 0
                    incorrect_ac_num = []
                    detection_pt_info = {}
                    
                    image_dir_name = "accumulator_sigma_" + str(std_dev) + "_edge_sigma_" + str(std_dev_edges) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD)
                    
                    print("Currently on: " + image_dir_name)
                    
                    if os.path.isdir(image_dir_name) != 1:
                        os.mkdir(image_dir_name)
                    else:
                        continue
                    
                    
                    #Get the ac_num to put into multi processing
                    multi_proc_ac_num = []
                    
                    for ac_num in ac_nums[0:10]:
                        if ac_num in ground_truth.keys():
                            multi_proc_ac_num.append(ac_num)
                    
                    print(multi_proc_ac_num)
                    
                    #Get optimal points through multi processing
                    with Pool(20) as p:
                        #print(std_dev)
                        #print(std_dev_edges)
                        #print(std_dev_canny)
                        #print(MIN_CANNY_THRESHOLD)
                        #print(MAX_CANNY_THRESHOLD)
                        
                        optimal_pts = p.map(functions.GHT,multi_proc_ac_num)
                        print(optimal_pts)
                    
                    
                    #Go through GHT for the validation set
                    for ac_num in ac_nums[0:10]:
                        if ac_num in ground_truth.keys():
                            #image_file_name = ac_num + "_accumulator_sigma_" + str(std_dev) + "_edge_sigma_" + str(std_dev_edges)  + "_canny_sigma_" + str(std_dev_canny) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD)
                            #total_detections = total_detections + 1
                            
                            
                            #optimal_pt = GHT(ac_num)
                            
                            optimal_pt = optimal_pts[0]
                            
                            print("Detected Optimal Point: ", optimal_pt)
                            print("Ground Truth Point: ", ground_truth[ac_num])
                        
                            curr_error = abs(np.linalg.norm(np.subtract(optimal_pt,ground_truth[ac_num])))**2 
                            error = error + curr_error
                            
                            #Can adjust threshold for correct detection accordingly
                            #if curr_error <= detection_threshold:
                            if curr_error <= 12.0:
                                correct_detections = correct_detections + 1
                            else:
                                incorrect_ac_num.append(ac_num)
                                
                            #Keep record of the information
                            detection_pt_info[ac_num] = [optimal_pt, ground_truth[ac_num], curr_error]
                    
                    plt.close()
                    
                
                    
                    print("======================================")
                    print("********SUMMARY OF PERFORMANCE********")
                    print("======================================")
                
                    print("Min Canny Threshold: ", MIN_CANNY_THRESHOLD)
                    print("Max Canny Threshold: ", MAX_CANNY_THRESHOLD)
                    print("Sigma Canny: ", std_dev_canny)
                    print("Sigma Accumulator: ", std_dev)
                    print("Sigma Edges: ", std_dev_edges)
                
                    print("The squared error for this trial on the validation set is :", error)
                    print("The detection rate is: " + str(correct_detections) + "/" + str(len(ground_truth.keys())))
                    
                    print("The Accession Numbers for Incorrect Detections are: ", incorrect_ac_num)
                    print("Detection Point Information: ", detection_pt_info)
                    
                    #Output General Information to File
                    f = open(image_dir_name + "/summary.txt","w")
                    f.write("======================================\n")
                    f.write("********SUMMARY OF PERFORMANCE********\n")
                    f.write("======================================\n")
                    f.write("Min Canny Threshold: %s \n" % str(MIN_CANNY_THRESHOLD))
                    f.write("Max Canny Threshold: %s \n" % str(MAX_CANNY_THRESHOLD))
                    f.write("Sigma Canny: %s \n" % str(std_dev_canny))
                    f.write("Sigma Accumulator: %s \n" % str(std_dev))
                    f.write("Sigma Edges: %s \n\n" % str(std_dev_edges))
                    
                    f.write("The squared error for this trial on the validation set is: %s \n\n" % str(error))
                    
                    f.write("The number of correct detections is %s" % correct_detections)
                    f.write("/%s \n\n" % str(len(ground_truth.keys())))
                    
                    f.write("Incorrect Accession Numbers: \n")
                    for item in incorrect_ac_num:
                        f.write("%s  " % str(item))

                    f.write("\n\n")
                    f.write("Below are the detections points: \n")
                    for key in detection_pt_info.keys():
                        f.write("AC Num: %s    Detected Point: " % str(key))
                        
                        info = detection_pt_info[key]
                        f.write("%s    Actual Point: " % str(info[0]))
                        f.write("%s    Error: " %str(info[1]))
                        f.write("%s" % info[2])
                        f.write("\n")
                    
                    f.close()
                    
                    
#===================================================================================================
#===================================================================================================