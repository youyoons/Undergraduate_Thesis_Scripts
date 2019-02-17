import openpyxl
import os
import matplotlib.pyplot as plt
import numpy as np
import math

#os.chdir("C:\\Users\\yoons\\Documents\\4th Year Semester 2\\ESC499 - Thesis\\Undergraduate_Thesis_Scripts\\GHT")
os.chdir("C:\\Users\\yoons\\Documents\\ESC499\\Undergraduate_Thesis_Scripts\\GHT")

book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts_validation_set.xlsx")
sheet = book.active
row_count = sheet.max_row

ground_truth = {}

x_total = 0
y_total = 0
z_total = 0
counter = 0

for i in range(3,row_count+1): #divided by 3 as a test 
    
    ac_num_loc = sheet.cell(row = i,column = 1)
    ac_num = str(ac_num_loc.value)
    
    x = sheet.cell(row = i, column = 2).value
    y = sheet.cell(row = i, column = 3).value
    z = sheet.cell(row = i, column = 4).value
    print(x,y,z)

    
    if (x != None) and (y != None) and (z != None):
        ground_truth[ac_num] = [x,y,z]
    
        x_total = x_total + x
        y_total = y_total + y
        z_total = z_total + z
        counter = counter + 1

print(counter)
#Plot non-maximal suppression points
#*******************************************************************************
nms_x_pts = [] 
nms_y_pts = []

for key in ground_truth.keys():
    nms_x_pts.append(ground_truth[key][0])
    nms_y_pts.append(ground_truth[key][1])  

plt.title("Ground Truth Points Distribution")
plt.ylim(128,0)
plt.xlim(0,128)
plt.xlabel('y')
plt.ylabel('x')

plt.scatter(nms_y_pts,nms_x_pts, marker='o', color='g')

plt.scatter(y_total/counter, x_total/counter, marker= 'X', color = 'r')


#Plot the Prior Distribution Boundary
#*******************************************************************************
#Power of ellipse-shape
pwr = 4

#Right half of prior distribution limit
ellipse_x1 = np.zeros(59)
ellipse_y1 = np.zeros(59)

for i in range(59):
    ellipse_x1[i] = i
    ellipse_y1[i] = math.pow(34**pwr - (34*(i-29)/29)**pwr, math.pow(pwr,-1)) + 51
    
#plt.plot(ellipse_y1, ellipse_x1, marker = 'X', color = 'b')

#Left half of prior distribution limit
ellipse_x2 = np.zeros(59)
ellipse_y2 = np.zeros(59)

for i in range(59):
    ellipse_x2[i] = i
    ellipse_y2[i] = -math.pow(34**pwr - (34*(i-29)/29)**pwr, math.pow(pwr,-1)) + 51
    
#plt.plot(ellipse_y2, ellipse_x2, marker = 'X', color = 'b')

'''
x_mu = 29
x_sig = 5.348

y_mu = 34
y_sig = 8.752

x, y = np.mgrid[0:58,0:68]
x_pwr = (x - x_mu)**2/(2*x_sig**2)
y_pwr = (y - y_mu)**2/(2*y_sig**2)

g = np.exp(-(x_pwr+y_pwr))

h = np.ones((58,85))
h[:,17:85] = np.log(g)
plt.gray()
plt.imshow(h)
'''

x1 = 0
x2 = 58
y1 = 17
y2 = 85

prior = np.zeros((x2-x1,y2-y1))

#Using Prior Distribution (about average centre of Ground Truth Points)
#Centre Variables
x_c = (x2+x1)//2
y_c = (y2+y1)//2

#Width
x_w = (x2-x1)//2
y_w = (y2-y1)//2


pwr = 2

#Just use widths as we are starting from the hard cut-out region
for dim1 in range(x2-x1):
    for dim2 in range(y2-y1):
        if (float(dim1-x_w)/x_w)**pwr + (float(dim2 - y_w)/y_w)**pwr <= 1:
            prior[dim1][dim2] = math.pow(1 - (float(dim1 - x_w)/x_w)**pwr - (float(dim2 - y_w)/y_w)**pwr,math.pow(pwr,-1))


h = np.ones((58,85))
h[:,17:85] = prior
plt.gray()
plt.imshow(h)



#Plot the hard bounding box
#*******************************************************************************
bound_sq_x = [0, 0, 58 , 58, 0]
bound_sq_y = [17, 85, 85, 17, 17]

plt.plot(bound_sq_y, bound_sq_x, marker = '.', color = 'k')


#Save Float and Show it
#*******************************************************************************
plt.savefig("ground_truth_distribution.png")
plt.show()

print(x_total/counter,y_total/counter,z_total/counter)