import os
import numpy as np
import math
import matplotlib.pyplot as plt
from collections import defaultdict
from scipy.misc import imread
from skimage.feature import canny
from scipy.ndimage.filters import sobel
from mpl_toolkits.mplot3d import Axes3D
import cv2 as cv
from scipy import signal
from scipy.ndimage.filters import gaussian_filter
import datetime
import openpyxl
import random
import cython
from multiprocessing import Pool

try:
    import cPickle
except ImportError:
    import pickle as cPickle


def gradient_orientation(scan):
    '''
    Calculate the gradient orientation for edge point in the scan
    '''
    #scipy.ndimage.sobel
    dx = sobel(scan, axis=0, mode='constant')	
    dy = sobel(scan, axis=1, mode='constant')
    dz = sobel(scan, axis=1, mode='constant')
    
    #For 3D instead of a single gradient value, we need two angles that define a normal vector
    #Phi is the angle between the positive x-axis to the projection of the normal vector the x-y plane (around +z)
    #Psi is the angle between the positive z-axis to the normal vector
    
    phi = np.arctan2(dy ,dx) * 180 / np.pi
    psi = np.arctan2(np.sqrt(dx*dx + dy*dy), dz) * 180 / np.pi
    

    gradient = np.zeros(scan.shape)
    
    return phi, psi

def build_r_table(referenceImage, origin):
    '''
    Build the R-table from the given referenceImage and a reference point
    '''
    edges = canny_edges_3d(referenceImage)
    
    phi, psi = gradient_orientation(edges)
    
    r_table = defaultdict(list)
    for (i,j,k),value in np.ndenumerate(edges):
        if value:
            r_table[(int(phi[i,j,k]),int(psi[i,j,k]))].append((origin[0]-i, origin[1]-j, origin[2] - k))
    
    
    return r_table

def canny_edges_3d(scan):
    dim = np.shape(scan)
    
    edges_x = np.zeros(scan.shape, dtype=bool) 
    edges_y = np.zeros(scan.shape, dtype=bool) 
    edges_z = np.zeros(scan.shape, dtype=bool) 
    edges = np.zeros(scan.shape, dtype=bool) 
    

    for i in range(dim[0]):
        edges_x[i,:,:] = canny(scan[i,:,:], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = std_dev_canny)
   
    for j in range(dim[1]):
        edges_y[:,j,:] = canny(scan[:,j,:], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = std_dev_canny)
        
    for k in range(dim[2]):
        edges_z[:,:,k] = canny(scan[:,:,k], low_threshold=MIN_CANNY_THRESHOLD, high_threshold=MAX_CANNY_THRESHOLD, sigma = std_dev_canny)
    
    
    for i in range(dim[0]):
        for j in range(dim[1]):
            for k in range(dim[2]):
                edges[i,j,k] = (edges_x[i,j,k]) or (edges_y[i,j,k]) or (edges_z[i,j,k])
    
    
    return edges


def accumulate_gradients(r_table, queryImage):
    '''
    Perform a General Hough Transform with the given image and R-table
    '''
    
    #Get edges matrix from Canny 
    edges = canny_edges_3d(queryImage) 
    
    #Get gradient angles
    phi, psi = gradient_orientation(edges)
    
    accumulator = np.zeros(queryImage.shape)
    accum_i = 0
    accum_j = 0 
    accum_k = 0

    edges_dim = np.shape(edges)

    #print(datetime.datetime.now())
    
    
    for (i,j,k),value in np.ndenumerate(edges):
        if value: 
            for r in r_table[(int(phi[i,j,k]), int(psi[i,j,k]))]:
                #iterations = iterations + 1
                accum_i, accum_j, accum_k = i+r[0], j+r[1], k+r[2]
                if accum_i < accumulator.shape[0] and accum_j < accumulator.shape[1] and accum_k < accumulator.shape[2]:
                    accumulator[int(accum_i), int(accum_j), int(accum_k)] += 1 
                    
    #print(datetime.datetime.now())  

    return accumulator

def general_hough_closure(reference_image):
    '''
    Generator function to create a closure with the reference image and origin
    at the center of the reference image
    
    Returns a function f, which takes a query image and returns the accumulator
    '''
    
    referencePoint = (reference_image.shape[0]/2, reference_image.shape[1]/2, reference_image.shape[2]/2)
    
    r_table = build_r_table(reference_image, referencePoint)
    
    def f(query_image):
        return accumulate_gradients(r_table, query_image)
        
    return f

def n_max(a, n):
    '''
    Return the N max elements and indices in a list
    '''
    indices = (-a.ravel()).argsort()[:n]
    indices = (np.unravel_index(i, a.shape) for i in indices)
    return [(a[i], i) for i in indices]
 
def test_general_hough(gh, reference_image, query):
    '''
    Uses a GH closure to detect shapes in an image and create nice output
    '''
    query_image = query
    query_dim = np.shape(query)
    reference_dim = np.shape(reference_image)

    accumulator = gh(query_image)

    return accumulator


#===================================================================================================
#****************************************** START OF GHT *******************************************
#===================================================================================================
def GHT(ac_num):
#===================================================================================================
#Obtaining the list of references to use
#===================================================================================================
    reference_acs = []

    image_file_name = ac_num + "_accumulator_sigma_" + str(std_dev) + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD) + "_prior_" + prior_type

    for file_name in os.listdir("no_fractures"):
        #The name should have "reference" in it but no "edge" in it
        if file_name.find("reference.pkl") != -1 and file_name.find("edge") == -1:
            #The reference list should not include the reference from the current ac_num
            if file_name.find(str(ac_num)) == -1:
                reference_acs.append(file_name)
    
#===================================================================================================
#Opening the Pickle File
#===================================================================================================    
    print("Accession Number: ", ac_num)
    
    #Open Downsized Pickle File Containing DICOM Scan
    try:
        dicom_dwn4x_pp = cPickle.load(open("no_fractures/dicom_3d_" + ac_num + "_dwn4x.pkl","rb"),encoding = 'latin1')
    except:
        dicom_dwn4x_pp = cPickle.load(open("no_fractures/dicom_3d_" + ac_num + "_dwn4x.pkl","rb"))
    
    dicom_dwn4x_pp_dim = np.shape(dicom_dwn4x_pp)
    print("Size of Downsized Dicom Input: ", dicom_dwn4x_pp_dim)


#===================================================================================================
#Use Prior to get a hard cutoff region
#===================================================================================================
    #Specify Region of Interest (Hard Blocking of Region Based on Prior Information)
    x1 = 0
    x2 = 60
    y1 = 17
    y2 = 87
    
    #Get specific region of focus (based on prior information)
    dicom_dwn4x = dicom_dwn4x_pp[x1:x2,y1:y2,:] #dicom_dwn4x contains the specific region of focus
    dicom_dwn4x_dim = np.shape(dicom_dwn4x)
    print("Size of Relevant Dicom (Prior Info): ", dicom_dwn4x_dim)
   

#===================================================================================================
#Obtain Final Accumulator Matrix through Max-Pooling of Individual Accumulator Matrices
#===================================================================================================
    #Initialize Accumulator that will be used to get top points
    accumulator = np.zeros(dicom_dwn4x_dim)
    
    #Choose N number of references
    random_reference_acs = reference_acs[:]

    #Set up dictionary to store temp accumulator matrices
    temp_accumulator_dict = {}

    for reference_ac in random_reference_acs:
        print("Current Reference: ", reference_ac)
        
        #Open up the Reference that is used as the reference image
        try:
            reference = cPickle.load(open("no_fractures/" + reference_ac,"rb"),encoding = 'latin1')
        except:
            reference = cPickle.load(open("no_fractures/" + reference_ac,"rb"))

        detect_s = general_hough_closure(reference)
        
        #Use max pooling on accumulator matrix
        temp_accumulator = test_general_hough(detect_s, reference, dicom_dwn4x)
        
        temp_accumulator_dict[reference_ac] = temp_accumulator
        
        accumulator = np.maximum(accumulator,temp_accumulator)
        
    #Set final_accumulator matrix
    final_accumulator = accumulator
    
    #Choose Top 3 References that are Most Relevant to Final Accumulator Matrix
    similarity_index = -float('inf')
    most_similar_reference_ac = None
    
    for reference_ac in temp_accumulator_dict.keys():
        temp_similarity_index = (temp_accumulator_dict[reference_ac] == final_accumulator).sum()
        #print("Reference AC")
        #print(reference_ac, temp_similarity_index)

        if temp_similarity_index > similarity_index:
            similarity_index = temp_similarity_index
            most_similar_reference_ac = reference_ac
        
    #Set most similar reference to plot later
    try:
        most_similar_reference = cPickle.load(open("no_fractures/" + most_similar_reference_ac,"rb"),encoding = 'latin1')
    except:
        most_similar_reference = cPickle.load(open("no_fractures/" + most_similar_reference_ac,"rb"))

    print("The most similar Reference is: ", most_similar_reference_ac)
        
    
    
    '''    
    plt.gray()
    plt.imshow(final_accumulator[:,:,ground_truth[ac_num][2]])
    plt.savefig("before_prior.png")
    '''
#===================================================================================================
#Using Prior Distribution
#===================================================================================================    
    #The final accumulator is the likelihood of the detection point being somewhere.
    #The prior is the function function: prior = (1 - (x-29)^4/29^4 - (y-51)^4/34^4)^1/4
    final_ac_dim = np.shape(final_accumulator)
    prior = np.zeros((final_ac_dim[0],final_ac_dim[1]))
    
    print(final_ac_dim)
    
    #Modify these parameters
    alpha = 1
    
    if prior_type == "ell_p4":
        pwr = 4
        
        for dim1 in range(final_ac_dim[0]):
            for dim2 in range(final_ac_dim[1]):
                if (float(dim1-29)/29)**pwr + (float(dim2 - 34)/34)**pwr <= 1:
                    prior[dim1][dim2] = math.pow(1 - (float(dim1 - 29)/29)**pwr - (float(dim2 - 34)/34)**pwr,math.pow(pwr,-1))
                    #print(math.pow(1 - (float(dim1 - 29)/29)**pwr - (float(dim2 - 34)/34)**pwr,math.pow(pwr,-1)))
                    #print(prior[dim1][dim2])
                    
        for dim3 in range(final_ac_dim[2]):
            #print("Prior")
            #print(prior[0:15,4:10])	 
            #print("Likelihood")
            #print(final_accumulator[0:15,4:10,dim3])
            
            #final_accumulator[:,:,dim3] = np.multiply(final_accumulator[:,:,dim3],prior)
            final_accumulator[:,:,dim3] = final_accumulator[:,:,dim3]*prior
            
            #print("Posterior")
            #print(final_accumulator[0:15,4:10,dim3])
            #print(likelihood)
            #print(final_accumulator[0:15,4:10,dim3])
    
    elif prior_type == "gaussian":
        x_mu = 29.5
        x_sig = 5.250
        
        y_mu = 34.5
        y_sig = 8.708
        
        x, y = np.mgrid[0:58,0:68]
        x_pwr = (x - x_mu)**2/(2*x_sig**2)
        y_pwr = (y - y_mu)**2/(2*y_sig**2)
        
        prior_pp = np.exp(-(x_pwr+y_pwr))
        
        #prior = np.log(prior_pp)
        prior = prior_pp
    
        for dim3 in range(final_ac_dim[2]):
            final_accumulator[:,:,dim3] = final_accumulator[:,:,dim3] + alpha*prior
    
    elif prior_type == "empirical":
        print("EMPIRICAL*************************")
        #print(final_accumulator[25:35,30:40,10])
        print(x_hist)
        print(y_hist)
        
        for dim1 in range(final_ac_dim[0]):
            for dim2 in range(final_ac_dim[1]):
                x_key = dim1//5
                y_key = dim2//5
                
                if (x_key in x_hist.keys()) and (y_key in y_hist.keys()):
                    prior_val = x_hist[x_key]*y_hist[y_key]
                else:
                    prior_val = 0
                
                prior[dim1][dim2] = prior_val
    
        for dim3 in range(final_ac_dim[2]):
            final_accumulator[:,:,dim3] = final_accumulator[:,:,dim3] + 100*prior
    
    else:
        pass
    
    #print(final_accumulator[25:35,30:40,10])
    
    
    plt.gray()
    plt.imshow(prior)
    plt.savefig("prior_empirical.png")
    plt.close()
    
            
#===================================================================================================
#Blurring the Accumulator Matrix and Query Edge Image
#===================================================================================================
    #Blur the final accumulator matrix
    final_accumulator = gaussian_filter(final_accumulator,sigma = std_dev, order = 0)

    #Blur the edge image for the whole dwn4x
    query_edges = canny_edges_3d(dicom_dwn4x_pp)

    query_edges_dim = np.shape(query_edges)
    
    query_edges_blurred = gaussian_filter(np.multiply(query_edges,50),sigma = std_dev_edges, order = 0)


#===================================================================================================
#Initial Plots and Top 40 Points for Visualization Purposes
#===================================================================================================
    plot_x = ground_truth[ac_num][0]
    plot_y = ground_truth[ac_num][1]
    plot_z = ground_truth[ac_num][2]


#===================================================================================================
#Get top 40 points
#===================================================================================================
    m = n_max(final_accumulator, 40)

    points = []
    x_pts = [] 
    y_pts = []
    z_pts = []
    
    for pt in m:
        points.append((pt[1][0] + x1,pt[1][1] + y1,pt[1][2], int(pt[0])))
    
        x_pts.append(pt[1][0]+x1)
        y_pts.append(pt[1][1]+y1) 
        z_pts.append(pt[1][2])


#===================================================================================================
#Take the top K average 
#===================================================================================================
    print("Top 5 Average")
    k = 5
    k_sum_pp = np.zeros(3)
    for index in range(k):
        k_sum_pp = np.add(k_sum_pp, m[index][1])
        print(m[index])
    
    optimal_pt = (int(k_sum_pp[0]//k) + x1,int(k_sum_pp[1]//k) + y1,int(k_sum_pp[2]//k))
    print("The optimal point is: ", optimal_pt)
    
#===================================================================================================
#Get the Non-maximal Suppression Points
#===================================================================================================
    #Perform non-maximal suppression
    nms_pts = []
    
    for pt in points:
        if len(nms_pts) == 0:
            nms_pts.append(pt)
        else:
            counter = 0
            for i in range(len(nms_pts)):
                if math.sqrt((nms_pts[i][0]-pt[0])**2 + (nms_pts[i][1]-pt[1])**2 + (nms_pts[i][2]-pt[2])**2) > 10:
                    counter = counter + 1
                else:
                    if pt[3] > nms_pts[i][3]:
                        nms_pts[i] = pt
            
            if counter == len(nms_pts):
                nms_pts.append(pt)
        
    print("Non-Maximal Suppression Points: ", nms_pts)
  

    #Original Optimal Point
    '''
    optimal_pt = [0,0,0]
    min_xdir = float('Inf')
 

    for pt in nms_pts:
        if pt[0] < min_xdir:
            min_xdir = pt[0]
            optimal_pt = pt[0:3]        
    '''  
    
    print("The Final Detection point is: ",optimal_pt)
    
     

    #Plot non-maximal suppression points
    nms_x_pts = [] 
    nms_y_pts = []
    nms_z_pts = []
    
    for pt in nms_pts:
        nms_x_pts.append(pt[0])
        nms_y_pts.append(pt[1]) 
        nms_z_pts.append(pt[2])    


#===================================================================================================
#Set up 3x3 Plot and put Query Image, Edge Image, Accumulator Slice
#=================================================================================================== 
    fig = plt.figure(num = image_file_name, figsize = (24,18))
    plt.gray()

    fig.suptitle(image_file_name)

    fig.add_subplot(4,3,1)
    plt.title('Query Image [Ground Truth Point: (' + str(plot_x) + ', ' + str(plot_y) + ', ' + str(plot_z) + ')]')
    plt.imshow(dicom_dwn4x_pp[:,:,plot_z])
    
    fig.add_subplot(4,3,2)
    plt.title('Query Image Edges')
    plt.imshow(query_edges[:,:,plot_z])
    
    fig.add_subplot(4,3,3)
    plt.title('Final Accumulator')
    plt.imshow(final_accumulator[:,:,plot_z])


#===================================================================================================
#Plot Top 40 Points
#===================================================================================================   
    #Sagittal View
    fig.add_subplot(4,3,4)
    plt.title('Top 40 Points (Sagittal)')
    plt.imshow(dicom_dwn4x_pp[:,:,plot_z])
    plt.scatter(y_pts,x_pts, marker='.', color='g')

    #Coronal View
    fig.add_subplot(4,3,5)
    plt.title('Top 40 Points (Coronal)')
    plt.imshow(dicom_dwn4x_pp[:,plot_y,:])
    plt.scatter(z_pts,x_pts, marker='.', color = 'g')
    
    #Axial View
    fig.add_subplot(4,3,6)
    plt.title('Top 40 Points (Axial)')
    plt.imshow(dicom_dwn4x_pp[plot_x,:,:])
    plt.scatter(z_pts,y_pts, marker='.', color = 'g')
    


#===================================================================================================
#Plot Detected (Optimal) and Ground Truth Points
#===================================================================================================
    #Sagittal View
    fig.add_subplot(4,3,7)
    plt.title('(Sagittal View) [Detected Point: (' + str(optimal_pt[0]) + ', ' + str(optimal_pt[1]) + ', ' + str(optimal_pt[2]) + ')]')
    plt.imshow(dicom_dwn4x_pp[:,:,plot_z])
    #plt.scatter(nms_y_pts,nms_x_pts, marker='o', color='g')
    plt.scatter(ground_truth[ac_num][1], ground_truth[ac_num][0], marker= 'o', color = 'y')
    plt.scatter(optimal_pt[1],optimal_pt[0], marker='X', color='r')
    
    #Coronal View
    fig.add_subplot(4,3,8)
    plt.title('Detected Point (Coronal View)')
    plt.imshow(dicom_dwn4x_pp[:,plot_y,:])
    plt.scatter(ground_truth[ac_num][2], ground_truth[ac_num][0], marker= 'o', color = 'y')
    plt.scatter(optimal_pt[2],optimal_pt[0], marker='X', color='r')
    
    #Axial View
    fig.add_subplot(4,3,9)
    plt.title('Detected Point (Axial View)')
    plt.imshow(dicom_dwn4x_pp[plot_x,:,:])
    plt.scatter(ground_truth[ac_num][2], ground_truth[ac_num][1], marker= 'o', color = 'y')
    plt.scatter(optimal_pt[2],optimal_pt[1], marker='X', color='r')
    

#===================================================================================================
#Plot Most Influential References
#===================================================================================================
    fig.add_subplot(4,3,10)
    plt.title('Most Influential Reference')
    plt.imshow(most_similar_reference[:,:,np.shape(most_similar_reference)[2]//2])

    
    #Save Figure
    print("Current Image: " + image_dir_name + "/" + image_file_name + ".png")
    plt.savefig(image_dir_name + "/" + image_file_name + ".png")
    
    
#===================================================================================================
#Add accumulator values around detected point to list in order to get a distribution
#===================================================================================================
    print(optimal_pt)
    for i in range(-2,3):
        for j in range(-2,3):
            for k in range(-2,3):  
                #print(final_accumulator[optimal_pt[0]-x1+i,optimal_pt[1]-y1+j,optimal_pt[2]+k])
                accum_dist.append(final_accumulator[optimal_pt[0]-x1+i,optimal_pt[1]-y1+j,optimal_pt[2]+k])

    #print(np.shape(final_accumulator))
    
    arr = np.array(final_accumulator)
    arr_flatten = arr.flatten()
    
    global total_accum
    total_accum = np.concatenate((total_accum, arr_flatten))
    '''
    try:
        total_accum = total_accum + arr.flatten()
    except:
        total_accum.append(arr.flatten())
    '''
    print(np.shape(total_accum))
    
    
    #Return the Detected Point
    return optimal_pt

#===================================================================================================
#******************************************* END OF GHT ********************************************
#===================================================================================================


#===================================================================================================
#===================================================================================================
if __name__ == '__main__':
    global prior_type
    prior_type = "empirical"
    mp = False 	
    
    if mp:
        os.chdir("../DicomSubsampling")
    else:    
        #os.chdir("C:\\Users\\yoons\\Documents\\ESC499\\Undergraduate_Thesis_Scripts\\DicomSubsampling")
        os.chdir("C:\\Users\\yoons\\Documents\\4th Year Semester 2\\Undergraduate_Thesis_Scripts\\DicomSubsampling")


    plt.close()
#===================================================================================================
#Process the accession numbers that are present and put it into a list
#===================================================================================================
    ac_nums_pp = os.listdir("no_fractures/")
    ac_nums = []
    
    for ac_num_pp in ac_nums_pp:
        if "reference" not in ac_num_pp:
            str1 = ac_num_pp.split("dicom_3d_")[1]
            str2 = str1.split("_")[0]
            
            ac_nums.append(str2)

#===================================================================================================
#Read in ground truth values from the ground_truth_detection_pts.xlsx spreadsheet
#===================================================================================================
    #Get the detection results for the validation set
    book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts_validation_set.xlsx")
    #book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts_test_set.xlsx")
    sheet = book.active
    row_count = sheet.max_row
    
    ground_truth = {}

    for i in range(3,row_count+1): 
    #for i in range(3,row_count+1):
        ac_num_loc = sheet.cell(row = i,column = 1)
        ac_num = str(ac_num_loc.value)
        
        x = sheet.cell(row = i, column = 2).value
        y = sheet.cell(row = i, column = 3).value
        z = sheet.cell(row = i, column = 4).value
        
        if (x != None) and (y != None) and (z != None):
            ground_truth[ac_num] = [x,y,z]
    
    global x_hist
    global y_hist
    global accum_dist #distribution of accumulator matrix around detected points
    global total_accum
    accum_dist = []
    total_accum = []
    
    x_hist = {}
    y_hist = {}
    
    #Creating Histogram Regarding Prior to be used in GHT
    for gt_pt in ground_truth.values():
        #print(gt_pt)
        x_val = (gt_pt[0]-0)//5
        y_val = (gt_pt[1]-17)//5
        
        #print(x_val,y_val)
        
        if x_val not in x_hist.keys():
            x_hist[x_val] = 0.01
        else:
            x_hist[x_val] = x_hist[x_val] + 0.01 #There are 100 points
            
        if y_val not in y_hist.keys():
            y_hist[y_val] = 0.01
        else:
            y_hist[y_val] = y_hist[y_val] + 0.01
       
    
    print(x_hist)
    print(y_hist)
        
    
#===================================================================================================
#Compute Detection Points, compare with ground truth to get error and detection rate
#===================================================================================================
    global std_dev
    global std_dev_edges
    global MIN_CANNY_THRESHOLD
    global MAX_CANNY_THRESHOLD
    global std_dev_canny
    global image_file_name
    global image_dir_name
    
    #Set Hyperparameters to be validated with validation set
    std_devs = [1.0]
    std_devs_edges = [0]
    min_cannys = [40]
    max_cannys = [160]

    std_dev_canny = 0.5
    
    for std_dev in std_devs:
        for std_dev_edges in std_devs_edges:
            for MIN_CANNY_THRESHOLD in min_cannys:
                for MAX_CANNY_THRESHOLD in max_cannys:
                    error = 0
                    correct_detections = 0
                    incorrect_ac_num = []
                    detection_pt_info = {}
                    
                    
                    image_dir_name = "prior_" + prior_type + "_min_canny_" + str(MIN_CANNY_THRESHOLD) + "_max_canny_" + str(MAX_CANNY_THRESHOLD) + "_accumulator_sigma_" + str(std_dev) 
                    
                    print("Currently on: " + image_dir_name)
                    
                    if os.path.isdir(image_dir_name) != 1:
                        os.mkdir(image_dir_name)
                    else:
                        continue
                    
                    if mp == False:
                        #Go through GHT for the validation set
                        for ac_num in ac_nums:
                            if ac_num in ground_truth.keys():
                                
                                optimal_pt = GHT(ac_num)
                                
                                print("Detected Optimal Point: ", optimal_pt)
                                print("Ground Truth Point: ", ground_truth[ac_num])
                            
                                curr_error = abs(np.linalg.norm(np.subtract(optimal_pt,ground_truth[ac_num])))**2 
                                error = error + curr_error
                                
                                #Can adjust threshold for correct detection accordingly
                                if curr_error <= 20.0:
                                    correct_detections = correct_detections + 1
                                else:
                                    incorrect_ac_num.append(ac_num)
                                    
                                #Keep record of the information
                                detection_pt_info[ac_num] = [optimal_pt, ground_truth[ac_num], curr_error]
                    else:
                        #Get the ac_num to put into multi processing
                        multi_proc_ac_num = []
                        
                        for ac_num in ac_nums:
                            if ac_num in ground_truth.keys():
                                multi_proc_ac_num.append(ac_num)
                        
                        print(multi_proc_ac_num)
                        
                        #Get optimal points through multi processing
                        p = Pool(processes = 25)
                
                        optimal_pts = p.map(GHT,multi_proc_ac_num)
                        
                        optimal_pts_dict = {}
                        #Put into dictionary
                        for i in range(len(multi_proc_ac_num)):
                            optimal_pts_dict[multi_proc_ac_num[i]] = optimal_pts[i]
                        
                        print(optimal_pts_dict)
    
                        #Go through GHT for the validation set
                        for ac_num in ac_nums:
                            if ac_num in ground_truth.keys():
                                
                                optimal_pt = optimal_pts_dict[ac_num]
                                print("Detected Optimal Point: ", optimal_pt)
                                print("Ground Truth Point: ", ground_truth[ac_num])
                            
                                curr_error = abs(np.linalg.norm(np.subtract(optimal_pt,ground_truth[ac_num])))**2 
                                error = error + curr_error
                                
                                #Can adjust threshold for correct detection accordingly
                                if curr_error <= 20.0:
                                    correct_detections = correct_detections + 1
                                else:
                                    incorrect_ac_num.append(ac_num)
                                    
                                #Keep record of the information
                                detection_pt_info[ac_num] = [optimal_pt, ground_truth[ac_num], curr_error]
                        
                    
                    plt.close()
                    
                
                    print("======================================")
                    print("********SUMMARY OF PERFORMANCE********")
                    print("======================================")
                
                    print("Min Canny Threshold: ", MIN_CANNY_THRESHOLD)
                    print("Max Canny Threshold: ", MAX_CANNY_THRESHOLD)
                    print("Sigma Canny: ", std_dev_canny)
                    print("Sigma Accumulator: ", std_dev)
                    print("Sigma Edges: ", std_dev_edges)
                
                    print("The squared error for this trial on the validation set is :", error)
                    print("The detection rate is: " + str(correct_detections) + "/" + str(len(ground_truth.keys())))
                    
                    print("The Accession Numbers for Incorrect Detections are: ", incorrect_ac_num)
                    print("Detection Point Information: ", detection_pt_info)
                    
                    
                    #Output General Information to File
                    f = open(image_dir_name + "/summary.txt","w")
                    f.write("======================================\n")
                    f.write("********SUMMARY OF PERFORMANCE********\n")
                    f.write("======================================\n")
                    f.write("Min Canny Threshold: %s \n" % str(MIN_CANNY_THRESHOLD))
                    f.write("Max Canny Threshold: %s \n" % str(MAX_CANNY_THRESHOLD))
                    f.write("Sigma Canny: %s \n" % str(std_dev_canny))
                    f.write("Sigma Accumulator: %s \n" % str(std_dev))
                    f.write("Sigma Edges: %s \n\n" % str(std_dev_edges))
                    
                    f.write("The squared error for this trial on the validation set is: %s \n\n" % str(error))
                    
                    f.write("The number of correct detections is %s" % correct_detections)
                    f.write("/%s \n\n" % str(len(ground_truth.keys())))
                    
                    f.write("Incorrect Accession Numbers: \n")
                    for item in incorrect_ac_num:
                        f.write("%s  " % str(item))

                    f.write("\n\n")
                    f.write("Below are the detections points: \n")
                    for key in detection_pt_info.keys():
                        f.write("AC Num: %s    Detected Point: " % str(key))
                        
                        info = detection_pt_info[key]
                        f.write("%s    Actual Point: " % str(info[0]))
                        f.write("%s    Error: " %str(info[1]))
                        f.write("%s" % info[2])
                        f.write("\n")
                    
                    f.close()
                    
                    #Create Excel Spreadsheet with Detection Information
                    wb = openpyxl.Workbook()
                    dest_filename = '../GHT/detection_pts_trial_' + str(MIN_CANNY_THRESHOLD) + '_' + str(MAX_CANNY_THRESHOLD) + '_' + str(std_dev_canny) + '_' + str(std_dev) + '_' + str(std_dev_edges) + '.xlsx'


                    ws1 = wb.active
                    ws1.append(['ac_num','x','y','z'])
                    for key in detection_pt_info.keys():
                        detected_pt = detection_pt_info[key][0]
                        row = [key,detected_pt[0],detected_pt[1],detected_pt[2]]

                        ws1.append(row)
                    wb.save(dest_filename)              
                    
                    
                    #Print Info About Accum Distribution
                    accum_local_dist_plot = {}
                    accum_dist_plot = {}
                    
                    #print(len(accum_dist))
                    #print(accum_dist)
                    #print(np.shape(total_accum))

                    
                    for index in range(len(accum_dist)):
                        temp_ind = accum_dist[index]//5
                        
                        if temp_ind not in accum_local_dist_plot.keys():
                            accum_local_dist_plot[temp_ind] = 1
                        else:
                            accum_local_dist_plot[temp_ind] = accum_local_dist_plot[temp_ind] + 1
                    
                    arr2 = np.array(total_accum)
                    total_accum_flat = arr2.flatten()
                    print(np.shape(arr2))
                    print("SHAPPEEEE")
                    print(np.shape(total_accum_flat))

                    
                    for index in range(len(total_accum_flat)):
                        temp_ind = int(total_accum_flat[index])//5
                        
                        if temp_ind not in accum_dist_plot.keys():
                            accum_dist_plot[temp_ind] = 1
                        else:
                            accum_dist_plot[temp_ind] = accum_dist_plot[temp_ind] + 1
                    
                    #for key in accum_dist_plot.keys():
                    #    print(key)
                    
                    print("============")
                    #print(accum_dist_plot)
                    #print(accum_local_dist_plot)
                    
                    print("============")
                    accum_dist_hist = []
                    
                    #print(accum_dist_plot.keys())
                    #print(max(accum_dist_plot.keys()))
                    
                    for index in range(int(max(accum_dist_plot.keys()))):
                        if index in accum_dist_plot.keys() and index in accum_local_dist_plot.keys():
                            accum_prob = accum_local_dist_plot[index]/accum_dist_plot[index]
                            accum_dist_hist.append([index,accum_prob])
                        else:
                            accum_dist_hist.append([index,0])
                            
                    print(accum_dist_hist)
                            
                            
                    
                    
                    
                        
#===================================================================================================
#===================================================================================================
