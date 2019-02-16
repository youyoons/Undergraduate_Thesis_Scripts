import os
import openpyxl

os.chdir("C:\\Users\\yoons\\Documents\\4th Year Semester 2\\Undergraduate_Thesis_Scripts\\GHT")

valid_acnum = []
test_acnum = []

#Get list of accession numbers in the validation set
book = openpyxl.load_workbook("../GHT/ground_truth_detection_pts_validation_set.xlsx")
sheet = book.active
row_count = sheet.max_row

ground_truth = {}

for i in range(3,row_count): #divided by 3 as a test 
    
    ac_num_loc = sheet.cell(row = i,column = 1)
    valid_acnum.append(str(ac_num_loc.value))



#Get all accession number present in no_fracutres
ac_nums_pp = os.listdir("..\\DicomSubsampling\\no_fractures")

for ac_num_pp in ac_nums_pp:
    if "reference" not in ac_num_pp:
        str1 = ac_num_pp.split("dicom_3d_")[1]
        str2 = str1.split("_")[0]
        
        if str2 not in valid_acnum:
            test_acnum.append(str2)


#Print Results
print("Validatin Set Acession Numbers")
print(sorted(valid_acnum))

print("\n\nTest Set Accession Numbers")
print(sorted(test_acnum))

