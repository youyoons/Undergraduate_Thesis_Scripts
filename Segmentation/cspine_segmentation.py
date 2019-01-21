import os
import numpy as np
import math
import matplotlib.pyplot as plt
from collections import defaultdict
from scipy.misc import imread
from skimage.feature import canny
from scipy.ndimage.filters import sobel
from mpl_toolkits.mplot3d import Axes3D
import cv2 as cv
from scipy import signal
from scipy.ndimage.filters import gaussian_filter
import datetime
import openpyxl

try:
    import cPickle
except ImportError:
    import pickle as cPickle

def cspine_segment(ac_num, detected_pt):
    dwn_dicom_filename = "../DicomSubsampling/no_fractures/dicom_3d_" + ac_num + "_dwn4x.pkl"

    try:
        try:
            dicom_dwn4x = cPickle.load(open(dwn_dicom_filename,"rb"),encoding = 'latin1')
        except:
            dicom_dwn4x = cPickle.load(open(dwn_dicom_filename,"rb"))    
    
        dicom_dwn4x_dim = np.shape(dicom_dwn4x)    

        x = detected_pt[0]
        y = detected_pt[1]
        z = detected_pt[2]

        #Get x, y, z ranges so that segmentation can be done 
        if x <= 20:
            x1 = 0
            x2 = 40
        else:
            x1 = x - 20
            x2 = x + 20

        if y <= 16:
            y1 = 0
            y2 = 32
        else:
            y1 = y - 16
            y2 = y + 16

        if z <= 12:
            z1 = 0
            z2 = 24
        else:
            z1 = z - 12
            z2 = z + 12

        segmented_dicom = dicom_dwn4x[x1:x2,y1:y2,z1:z2]

        #Save segmented_dicom in pkl file
        if os.path.isdir('segmented_pkl') != 1:
            os.mkdir('segmented_pkl')                           

        segmented_dicom_pkl = 'segmented_pkl/dicom_3d_' + ac_num + '_segmented.pkl'
        cPickle.dump(segmented_dicom, open(segmented_dicom_pkl,"wb"),protocol = 2) 
    except:
        print("Error trying to get segmented cspine vertebrae for Accession Number: ", ac_num)    
        
    return

def visualize_segmentation(ac_num):
    filename = 'segmented_pkl/dicom_3d_' + ac_num + '_segmented.pkl' 
    try:
        cspine_segment = cPickle.load(open(filename,"rb"),encoding = 'latin1')
    except:
        cspine_segment = cPickle.load(open(filename,"rb")) 

    segment_dim = np.shape(cspine_segment)
    


    #Plotting the Segmentation Result
    #Frontal
    fig = plt.figure(figsize = (24,8))
    plt.gray()

    for i in range(1,6):    
        fig.add_subplot(1,5,i)
        plt.imshow(cspine_segment[:,:,int(segment_dim[2]/6*i)])
    
    #Side 
    fig2 = plt.figure(figsize = (20,10))
    plt.gray()

    for i in range(1,6):    
        fig2.add_subplot(1,5,i)
        plt.imshow(cspine_segment[:,int(segment_dim[1]/6*i),:])
    
    plt.show()

if __name__ == '__main__':

    book = openpyxl.load_workbook('../GHT/detection_pts_trial_30_140_0.5_1.0_0.xlsx')
    sheet = book.active
    row_count = sheet.max_row

    detected_points = {}

    for i in range(2,row_count): #divided by 3 as a test 
        
        ac_num_loc = sheet.cell(row = i,column = 1)
        ac_num = str(ac_num_loc.value)
        
        x = sheet.cell(row = i, column = 2).value
        y = sheet.cell(row = i, column = 3).value
        z = sheet.cell(row = i, column = 4).value
        
        detected_points[ac_num] = [x,y,z]
   
        #Segment each detected point and save a pkl file for it
        cspine_segment (ac_num,detected_points[ac_num])

        #print(ac_num,x,y,z)


    print(detected_points)

    #A sample visualization
    visualize_segmentation('5826444')
    visualize_segmentation('9020776')



