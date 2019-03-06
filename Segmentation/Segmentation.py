import os
import numpy as np
import math
import matplotlib.pyplot as plt
from collections import defaultdict
from scipy.misc import imread
from skimage.feature import canny
from scipy.ndimage.filters import sobel
from mpl_toolkits.mplot3d import Axes3D
import cv2 as cv
from scipy import signal
from scipy.ndimage.filters import gaussian_filter
import datetime
import openpyxl
from glob import glob
import pydicom

try:
    import cPickle
except ImportError:
    import pickle as cPickle

def cspine_segment(ac_num, detected_pt):
    try:      
        #Create DICOM
        base_series_path = "/home/youy/Documents/Spine/ProcessedData_y/"
        
        #Get all series that are processed
        series_paths = glob(base_series_path + "/*/")
        
        for potential_path in series_paths:
            #print("Potential Path: ", potential_path)
            if ac_num in potential_path:
                series_path = potential_path
        
        print(series_path)
    
        final_series_path = glob(series_path + "/*/")[0]

        print("Series Path is: ", final_series_path)

        dicom_fullsize = create_3d_dicom(final_series_path)
    
        dicom_dim = np.shape(dicom_fullsize)    

        x = detected_pt[0]
        y = detected_pt[1]
        z = detected_pt[2]

        #Get x, y, z ranges so that segmentation can be done (160 in x direction, 160 in y direction, 56 in z direction) 
        if x <= 80:
            x1 = 0
            x2 = 160
        else:
            x1 = x - 80
            x2 = x + 80

        if y <= 64:
            y1 = 0
            y2 = 160
        else:
            y1 = y - 64
            y2 = y + 96

        if z <= 28:
            z1 = 0
            z2 = 56
        else:
            z1 = z - 28
            z2 = z + 28

        segmented_dicom = dicom_fullsize[x1:x2,y1:y2,z1:z2]
        #Save segmented_dicom in pkl file (160,160,56)
        if os.path.isdir('segmented_pkl') != 1:
            os.mkdir('segmented_pkl')                           

        segmented_dicom_pkl = 'segmented_pkl/dicom_3d_' + ac_num + '_segmented.pkl'
        cPickle.dump(segmented_dicom, open(segmented_dicom_pkl,"wb"),protocol = 2) 
    except:
        print("Error trying to get segmented cspine vertebrae for Accession Number: ", ac_num)    
        
    return

def visualize_segmentation(ac_num):
    filename = 'segmented_pkl/dicom_3d_' + ac_num + '_segmented.pkl' 
    try:
        cspine_segment = cPickle.load(open(filename,"rb"),encoding = 'latin1')
    except:
        cspine_segment = cPickle.load(open(filename,"rb")) 

    segment_dim = np.shape(cspine_segment)
    
    print("SEGMENT DIMENSION**********************")
    print(segment_dim)

    #Plotting the Segmentation Result
    #Frontal
    fig = plt.figure(figsize = (60,5))
    plt.gray()

    for i in range(segment_dim[2]//4):  
        #print(2*i+1)  
        fig.add_subplot(1,segment_dim[2]//4,i+1)
        #plt.imshow(cspine_segment[:,:,int(segment_dim[2]/6*i)])
        plt.imshow(cspine_segment[:,:,4*i+1])
    
    #Side 
    fig2 = plt.figure(figsize = (60,5))
    plt.gray()

    for i in range(segment_dim[1]//8): 
        #print(2*i)   
        fig2.add_subplot(1,segment_dim[1]//8,i+1)
        plt.imshow(cspine_segment[:,i*8+1,:])
    
    plt.show()
    
    
#Key: Accession Number (Converted to String)
#Value: Path (String)
def create_dict_paths(is_fracture):
    #Get spreadsheet that contains paths to CSpine Accession Numbers
    book = openpyxl.load_workbook('../Copy_Scans/CS_accession_number_paths.xlsx')
    sheet = book.active
    row_count = sheet.max_row
    
    dict_paths = {}
    
    for row in range(2,row_count+1):
        #Fractures
        if is_fracture:
            
            if sheet.cell(row,3).value == 1 or sheet.cell(row,4).value == 1:
                ac_num = str(sheet.cell(row,1).value)
                path_scan = sheet.cell(row,2).value
            
                if (ac_num not in dict_paths.keys()) and (path_scan != None):  
                    dict_paths[ac_num]  = path_scan
        #No Fractures
        else:
            if sheet.cell(row,3).value != 1 and sheet.cell(row,4).value != 1:
                ac_num = str(sheet.cell(row,1).value)
                path_scan = sheet.cell(row,2).value
            
                if (ac_num not in dict_paths.keys()) and (path_scan != None):  
                    dict_paths[ac_num]  = path_scan
    
    return dict_paths


#Function: takes in the path of a series and returns a 3D Numpy Array
def create_3d_dicom(series_path):
    dicom_filenames = sorted(os.listdir(series_path))
    dicom_files = [file_name for file_name in dicom_filenames if file_name[-4:]==".dcm"]
    print(dicom_files)
    
    slice_num = 0

    ds = pydicom.dcmread(series_path + dicom_files[0])
    image_2d = ds.pixel_array.astype(float)
    size_2d = np.shape(image_2d)
    dicom_3d = np.zeros((size_2d[0],size_2d[1],len(dicom_files)))

    for file_name in dicom_files:
        ds = pydicom.dcmread(series_path + file_name)
    
        image_2d = ds.pixel_array.astype(float)
        image_2d_scaled = (np.maximum(image_2d,0) / image_2d.max()) * 255.0
    
        image_2d_scaled = np.uint8(image_2d_scaled)
        
        dicom_3d[:,:,slice_num] = image_2d_scaled
    
        slice_num += 1
    
    print(dicom_3d.shape)
    
    return dicom_3d

if __name__ == '__main__':
    '''
    frac_paths_dict = {}
    no_frac_paths_dict = {}
    
    frac_paths_dict = create_dict_paths(True)
    no_frac_paths_dict = create_dict_paths(False)
    
    #print(len(frac_paths_dict)) #274
    #print(len(no_frac_paths_dict)) #7898
    '''

    #Going through results from GHT
    book = openpyxl.load_workbook('../GHT/detection_pts_trial_40_160_0.5_1.0_0.xlsx')
    sheet = book.active
    row_count = sheet.max_row
    
    detected_points = {}
    
    for i in range(2,row_count+1): #divided by 3 as a test 
        
        ac_num_loc = sheet.cell(row = i,column = 1)
        ac_num = str(ac_num_loc.value)
        
        x = sheet.cell(row = i, column = 2).value
        y = sheet.cell(row = i, column = 3).value
        z = sheet.cell(row = i, column = 4).value
        
        #Modify Detection Point for Full-sized
        detected_points[ac_num] = [4*x,4*y,2*z] 
        
        
        #Segment each detected point and save a pkl file for it
        #cspine_segment(ac_num,detected_points[ac_num])


    print(detected_points)
    

    #A sample visualization
    visualize_segmentation('8864910')
    #visualize_segmentation('9020776')
    


