import openpyxl
import os
import numpy as np

#os.chdir('C:\\Users\\yoons\\Documents\\ESC499\\Undergraduate_Thesis_Scripts\\Copy_Scans')


#True if we want to get paths with only fractures
#False if we want to get paths without any fractures
fracture = False

#Get spreadsheet that contains paths to CSpine Accession Numbers
book = openpyxl.load_workbook('CS_accession_number_paths.xlsx')
sheet = book.active
row_count = sheet.max_row

scans_log = open("path_scans.log","w")


for row in range(2,row_count+1):
    #Fractures
    if fracture:
        if sheet.cell(row,3).value == 1 or sheet.cell(row,4).value == 1:
            path_scan = sheet.cell(row,2).value
        
            if path_scan != None:   
                scans_log.write(path_scan + "\n")
    #No Fractures
    else:
        if sheet.cell(row,3).value != 1 and sheet.cell(row,4).value != 1:
            path_scan = sheet.cell(row,2).value
        
        if path_scan != None:
            scans_log.write(path_scan + "\n")

scans_log.close()
